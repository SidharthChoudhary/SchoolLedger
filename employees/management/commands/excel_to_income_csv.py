"""
Management command — Step 1 of 2-step historical income import.

Reads BOTH income sheets from the school accounts Excel:
  • VIncF — Fee income vouchers (Tuition Fee, Bus Fee, RTE)
  • VIncO — Other income vouchers (loans, bus bookings, misc)

Produces a single CSV ready for upload at:
    /ledger-income/bulk-import-ledger/

Usage:
    python manage.py excel_to_income_csv --excel "path/to/file.xlsm"
    python manage.py excel_to_income_csv --excel "path/to/file.xlsm" --dry-run
    python manage.py excel_to_income_csv --excel "path/to/file.xlsm" --no-other   # skip VIncO
    python manage.py excel_to_income_csv --excel "path/to/file.xlsm" --no-fees    # skip VIncF
"""

import csv
import os

from django.core.management.base import BaseCommand, CommandError

from dailyLedger.models import Session

OUTPUT_FIELDS = [
    'Voucher_Number', 'Date', 'Amount',
    'Major_Head', 'Head', 'Sub_Head',
    'Payment_Type', 'Session', 'Details',
]

OUTPUT_DIR = 'datamigration/convertedcsvs'


def _session_from_date(dt):
    """datetime → '2024-2025' (April starts new session)"""
    if dt is None:
        return None
    yr = dt.year if dt.month >= 4 else dt.year - 1
    return f'{yr}-{yr + 1}'


# Fee type → (Major_Head, Head, Sub_Head) mapping
FEE_TYPE_MAP = {
    'tuition fee':   ('Income', 'Fees', 'Tuition Fee'),
    'bus fee':       ('Income', 'Fees', 'Bus Fee'),
    'rte':           ('Income', 'Fees', 'RTE'),
    'exam/tc fee':   ('Income', 'Fees', 'Exam/TC Fee'),
    'admission fee': ('Income', 'Fees', 'Admission Fee'),
}
FEE_TYPE_DEFAULT = ('Income', 'Fees', 'Tuition Fee')


class Command(BaseCommand):
    help = (
        'STEP 1 — Convert VIncF + VIncO sheets of the school accounts Excel into '
        'a CSV ready for /ledger-income/bulk-import-ledger/'
    )

    def add_arguments(self, parser):
        parser.add_argument('--excel', required=True,
                            help='Full path to the .xlsm / .xlsx workbook')
        parser.add_argument('--output',
                            default=os.path.join(OUTPUT_DIR, 'income_import.csv'),
                            help=f'Output CSV path (default: {OUTPUT_DIR}/income_import.csv)')
        parser.add_argument('--session', default=None,
                            help='Force a single session string, e.g. 2024-2025 (auto-detected)')
        parser.add_argument('--no-fees', action='store_true',
                            help='Skip VIncF (fee income) sheet')
        parser.add_argument('--no-other', action='store_true',
                            help='Skip VIncO (other income) sheet')
        parser.add_argument('--skip-zero', action='store_true',
                            help='Skip rows where Amount is 0')
        parser.add_argument('--dry-run', action='store_true',
                            help='Print summary without writing a file')

    def handle(self, *args, **options):
        excel_path    = options['excel']
        output_path   = options['output']
        force_session = options['session']
        skip_fees     = options['no_fees']
        skip_other    = options['no_other']
        skip_zero     = options['skip_zero']
        dry_run       = options['dry_run']

        if not os.path.exists(excel_path):
            raise CommandError(f'File not found: {excel_path}')

        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is not installed.  Run: pip install openpyxl')

        valid_sessions = {s.session for s in Session.objects.all()}

        import warnings
        warnings.filterwarnings('ignore')
        wb = openpyxl.load_workbook(excel_path, read_only=True, keep_vba=True)

        output_rows = []
        bad_session = []

        # ── VIncF — Fee income vouchers ────────────────────────────────────
        if not skip_fees:
            sheet_name = 'VIncF'
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet "{sheet_name}" not found — skipping.'))
            else:
                self.stdout.write(f'Reading {sheet_name}...')
                ws = wb[sheet_name]
                raw = list(ws.iter_rows(min_row=2, max_col=10, values_only=True))
                # Columns: Voucher#, Date, Amount, Acc Number, Acc Name, Class, Fathers Name, Fee Type, Remark
                vinc_f_rows = 0
                for row_num, row in enumerate(raw, start=2):
                    voucher_num = row[0]
                    date_val    = row[1]
                    amount      = row[2]
                    acc_name    = row[4]   # student name
                    fee_type    = row[7]
                    remark      = row[8]

                    if not date_val or not isinstance(date_val, __import__('datetime').datetime):
                        continue
                    if amount is None:
                        continue
                    try:
                        amount = float(amount)
                    except (ValueError, TypeError):
                        continue
                    if skip_zero and amount == 0:
                        continue

                    date_str = date_val.strftime('%Y-%m-%d')

                    session_str = force_session or _session_from_date(date_val)
                    if session_str not in valid_sessions:
                        bad_session.append((row_num, sheet_name, date_str, session_str))
                        session_str = ''

                    ft_key = str(fee_type).lower().strip() if fee_type else ''
                    major_head, head, sub_head = FEE_TYPE_MAP.get(ft_key, FEE_TYPE_DEFAULT)

                    # For fee income, Sub_Head = the specific fee type, Details = student name
                    details = str(acc_name).strip() if acc_name else ''
                    if remark and str(remark).strip():
                        details = (details + ' | ' + str(remark).strip()).strip(' |')
                    details = details[:200]

                    output_rows.append({
                        'Voucher_Number': str(voucher_num).strip() if voucher_num else '',
                        'Date':          date_str,
                        'Amount':        round(amount, 2),
                        'Major_Head':    major_head,
                        'Head':          head,
                        'Sub_Head':      sub_head,
                        'Payment_Type':  'Cash',
                        'Session':       session_str,
                        'Details':       details,
                    })
                    vinc_f_rows += 1

                self.stdout.write(self.style.SUCCESS(f'  VIncF rows: {vinc_f_rows}'))

        # ── VIncO — Other income vouchers ──────────────────────────────────
        if not skip_other:
            sheet_name = 'VIncO'
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet "{sheet_name}" not found — skipping.'))
            else:
                self.stdout.write(f'Reading {sheet_name}...')
                ws = wb[sheet_name]
                raw = list(ws.iter_rows(min_row=2, max_col=9, values_only=True))
                # Columns: VoucherNumber, Date, Amount, Remark, Name, Major Head, Head, Sub Head, Month(formula)
                vinc_o_rows = 0
                for row_num, row in enumerate(raw, start=2):
                    voucher_num = row[0]
                    date_val    = row[1]
                    amount      = row[2]
                    remark      = row[3]
                    name        = row[4]
                    major_head  = row[5]
                    head        = row[6]
                    sub_head    = row[7]

                    if not date_val or not isinstance(date_val, __import__('datetime').datetime):
                        continue
                    if amount is None:
                        continue
                    try:
                        amount = float(amount)
                    except (ValueError, TypeError):
                        continue
                    if skip_zero and amount == 0:
                        continue

                    date_str = date_val.strftime('%Y-%m-%d')

                    session_str = force_session or _session_from_date(date_val)
                    if session_str not in valid_sessions:
                        bad_session.append((row_num, sheet_name, date_str, session_str))
                        session_str = ''

                    # Most VIncO rows have no head classification — use "Other Income" defaults
                    eff_major = str(major_head).strip() if major_head else 'Income'
                    eff_head  = str(head).strip()       if head       else 'Other Income'
                    eff_sub   = str(sub_head).strip()   if sub_head   else (str(name).strip() if name else '')

                    # Derive Sub_Head from remark when still empty
                    if not eff_sub and remark:
                        remark_str = str(remark).strip().lower()
                        if 'loan' in remark_str:
                            eff_sub = 'Loan'
                        elif 'bus booking' in remark_str or 'bus' in remark_str:
                            eff_sub = 'Bus Booking'
                        elif 'sale' in remark_str:
                            eff_sub = 'Misc Sale'
                        else:
                            eff_sub = 'Miscellaneous'
                    elif not eff_sub:
                        eff_sub = 'Miscellaneous'

                    details = str(remark).strip() if remark else ''
                    details = details[:200]

                    remark_lower = details.lower()
                    if 'loan' in remark_lower:
                        payment_type = 'Credit'
                    elif 'bank' in remark_lower or 'neft' in remark_lower or 'upi' in remark_lower:
                        payment_type = 'Bank Transfer'
                    else:
                        payment_type = 'Cash'

                    output_rows.append({
                        'Voucher_Number': str(voucher_num).strip() if voucher_num else '',
                        'Date':          date_str,
                        'Amount':        round(amount, 2),
                        'Major_Head':    eff_major,
                        'Head':          eff_head,
                        'Sub_Head':      eff_sub,
                        'Payment_Type':  payment_type,
                        'Session':       session_str,
                        'Details':       details,
                    })
                    vinc_o_rows += 1

                self.stdout.write(self.style.SUCCESS(f'  VIncO rows: {vinc_o_rows}'))

        # ── Summary ────────────────────────────────────────────────────────
        self.stdout.write(self.style.SUCCESS(f'\nTotal rows to write: {len(output_rows)}'))

        if bad_session:
            self.stdout.write(self.style.WARNING(
                f'{len(bad_session)} rows had unrecognised session (left blank in CSV):'
            ))
            for row_num, sheet, d, s in bad_session[:10]:
                self.stdout.write(self.style.WARNING(f'  {sheet} Row {row_num}: date {d} → session "{s}"'))

        if dry_run:
            self.stdout.write(self.style.WARNING('\nDRY RUN — no file written.'))
            self._preview(output_rows[:8])
            return

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
            writer.writeheader()
            writer.writerows(output_rows)

        self.stdout.write(self.style.SUCCESS(f'\n[OK] Written to: {os.path.abspath(output_path)}'))
        self.stdout.write(
            '\nSTEP 2 — Upload at:\n'
            '  http://127.0.0.1:8000/ledger-income/bulk-import-ledger/'
        )

    def _preview(self, rows):
        if not rows:
            return
        self.stdout.write('\nPreview (first 8 rows):')
        for r in rows:
            self.stdout.write(str(r))
