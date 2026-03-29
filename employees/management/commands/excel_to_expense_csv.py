"""
Management command — Step 1 of 2-step historical expense import.

Reads the VExp sheet (expense vouchers) from the school accounts Excel and
produces a CSV ready for upload at:
    /ledger-expense/bulk-import-ledger/

Usage:
    python manage.py excel_to_expense_csv --excel "path/to/file.xlsm"
    python manage.py excel_to_expense_csv --excel "path/to/file.xlsm" --dry-run
"""

import csv
import os

from django.core.management.base import BaseCommand, CommandError

from employees.models import Employee
from dailyLedger.models import Session

OUTPUT_FIELDS = [
    'Voucher_Number', 'Date', 'Amount',
    'Major_Head', 'Head', 'Sub_Head',
    'Payment_Type', 'Session', 'Details',
]

OUTPUT_DIR = 'datamigration/convertedcsvs'

# How to infer the session string from a date
def _session_from_date(dt):
    """datetime → '2024-2025' (April starts new session)"""
    if dt is None:
        return None
    yr = dt.year if dt.month >= 4 else dt.year - 1
    return f'{yr}-{yr + 1}'


class Command(BaseCommand):
    help = (
        'STEP 1 — Convert VExp sheet of the school accounts Excel into a CSV '
        'ready for /ledger-expense/bulk-import-ledger/'
    )

    def add_arguments(self, parser):
        parser.add_argument('--excel', required=True,
                            help='Full path to the .xlsm / .xlsx workbook')
        parser.add_argument('--output',
                            default=os.path.join(OUTPUT_DIR, 'expense_import.csv'),
                            help=f'Output CSV path (default: {OUTPUT_DIR}/expense_import.csv)')
        parser.add_argument('--sheet', default='VExp',
                            help='Sheet name (default: VExp)')
        parser.add_argument('--session', default=None,
                            help='Force a single session string, e.g. 2024-2025 (auto-detected by default)')
        parser.add_argument('--skip-zero', action='store_true',
                            help='Skip rows where Amount is 0')
        parser.add_argument('--dry-run', action='store_true',
                            help='Print summary without writing a file')

    def handle(self, *args, **options):
        excel_path  = options['excel']
        output_path = options['output']
        sheet_name  = options['sheet']
        force_session = options['session']
        skip_zero   = options['skip_zero']
        dry_run     = options['dry_run']

        if not os.path.exists(excel_path):
            raise CommandError(f'File not found: {excel_path}')

        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is not installed.  Run: pip install openpyxl')

        # Pre-load valid sessions from DB for validation
        valid_sessions = {s.session for s in Session.objects.all()}
        # Build employee name → employee lookup for sub_head enrichment
        emp_names = {e.name.lower().strip() for e in Employee.objects.all()}

        self.stdout.write(f'Reading: {excel_path}  (sheet: {sheet_name})')

        import warnings
        warnings.filterwarnings('ignore')
        wb = openpyxl.load_workbook(excel_path, read_only=True, keep_vba=True)

        if sheet_name not in wb.sheetnames:
            raise CommandError(f'Sheet "{sheet_name}" not found.\nAvailable: {wb.sheetnames}')

        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(min_row=1, max_col=10, values_only=True))

        header = raw_rows[0]
        # Expected: VoucherNumber, Date, Amount, Remark, Name, Major Head, Head, Sub Head, Month
        col = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}

        output_rows  = []
        skipped      = []
        bad_session  = []

        for row_num, row in enumerate(raw_rows[1:], start=2):
            # Pull values by positional index (columns are fixed in this sheet)
            voucher_num = row[0]
            date_val    = row[1]
            amount      = row[2]
            remark      = row[3]      # also used as Details / Payment mode
            name        = row[4]      # employee name (for salary rows)
            major_head  = row[5]
            head        = row[6]
            sub_head    = row[7]

            # Skip blank / formula rows
            if not date_val or not isinstance(date_val, __import__('datetime').datetime):
                skipped.append(row_num)
                continue

            if amount is None:
                skipped.append(row_num)
                continue

            try:
                amount = float(amount)
            except (ValueError, TypeError):
                skipped.append(row_num)
                continue

            if skip_zero and amount == 0:
                skipped.append(row_num)
                continue

            # Format date
            date_str = date_val.strftime('%Y-%m-%d')

            # Determine session
            if force_session:
                session_str = force_session
            else:
                session_str = _session_from_date(date_val)

            if session_str not in valid_sessions:
                bad_session.append((row_num, date_str, session_str))
                session_str = ''       # leave blank — import will still work but without session link

            # Build sub_head: prefer explicit sub_head column; for salary rows use employee name
            effective_sub_head = ''
            if sub_head and str(sub_head).strip():
                effective_sub_head = str(sub_head).strip()
            if name and str(name).strip():
                # For salary entries the sub_head naturally equals the employee name
                effective_sub_head = str(name).strip()

            # Details: combine remark + name where both exist
            details_parts = [str(v).strip() for v in [remark] if v and str(v).strip()]
            details = ' | '.join(details_parts)[:200]

            # Payment type — infer from remark
            remark_lower = str(remark).lower() if remark else ''
            if 'bank' in remark_lower or 'neft' in remark_lower or 'upi' in remark_lower:
                payment_type = 'Bank Transfer'
            elif 'credit' in remark_lower:
                payment_type = 'Credit'
            else:
                payment_type = 'Cash'

            output_rows.append({
                'Voucher_Number': str(voucher_num).strip() if voucher_num else '',
                'Date':          date_str,
                'Amount':        round(amount, 2),
                'Major_Head':    str(major_head).strip() if major_head else '',
                'Head':          str(head).strip() if head else '',
                'Sub_Head':      effective_sub_head,
                'Payment_Type':  payment_type,
                'Session':       session_str,
                'Details':       details,
            })

        self.stdout.write(self.style.SUCCESS(f'Rows to write  : {len(output_rows)}'))
        self.stdout.write(f'Rows skipped   : {len(skipped)}')

        if bad_session:
            self.stdout.write(self.style.WARNING(
                f'{len(bad_session)} rows had unrecognised session (session left blank in CSV):'
            ))
            for row_num, d, s in bad_session[:10]:
                self.stdout.write(self.style.WARNING(f'  Row {row_num}: date {d} → session "{s}"'))

        if dry_run:
            self.stdout.write(self.style.WARNING('\nDRY RUN — no file written.'))
            self._preview(output_rows[:8])
            return

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
            writer.writeheader()
            writer.writerows(output_rows)

        self.stdout.write(self.style.SUCCESS(f'\n[OK] Written to: {os.path.abspath(output_path)}'))
        self.stdout.write(
            '\nSTEP 2 — Upload at:\n'
            '  http://127.0.0.1:8000/ledger-expense/bulk-import-ledger/'
        )

    def _preview(self, rows):
        if not rows:
            return
        self.stdout.write('\nPreview (first 8 rows):')
        for r in rows:
            self.stdout.write(str(r))
