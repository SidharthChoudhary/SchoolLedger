"""
Management command — Step 1 of the 2-step historical payroll import.

Usage:
    python manage.py excel_to_payroll_csv --excel "path/to/file.xlsm" --output payroll_import.csv

Step 2 (upload the generated CSV):
    Go to /employees/payroll/bulk-import/ and upload the output file.
"""

import csv
import os
import difflib

from django.core.management.base import BaseCommand, CommandError

from employees.models import Employee


# Ordered as they appear in the SdS sheet, April→March financial year
MONTH_ORDER = [
    'April', 'May', 'June', 'July', 'August', 'September',
    'October', 'November', 'December', 'January', 'February', 'March',
]

# Month name → month number
MONTH_NUM = {
    'April': 4, 'May': 5, 'June': 6, 'July': 7,
    'August': 8, 'September': 9, 'October': 10, 'November': 11,
    'December': 12, 'January': 1, 'February': 2, 'March': 3,
}

OUTPUT_FIELDS = [
    'Emp_ID', 'Month', 'Payable_Salary',
    'Old_Dues', 'Other_Amount', 'Note',
    'Manual_Work_Days', 'Manual_Leave_Days',
]


def _parse_session_years(session_str):
    """'2024-2025' → (2024, 2025)"""
    try:
        parts = str(session_str).split('-')
        return int(parts[0]), int(parts[1])
    except (ValueError, IndexError, TypeError):
        return None, None


def _month_str(month_name, start_year, end_year):
    """'April', 2024, 2025 → '2024-04'.  Jan/Feb/Mar use end_year."""
    mo = MONTH_NUM[month_name]
    yr = start_year if mo >= 4 else end_year
    return f'{yr}-{mo:02d}'


def _best_name_match(target, emp_lookup):
    """
    Find the best matching employee for a name string.
    Returns (Employee | None, match_type_str)
    """
    key = target.lower().strip()

    # 1. Exact match on name or display_name
    if key in emp_lookup['exact']:
        return emp_lookup['exact'][key], 'exact'

    # 2. Substring match: target contains emp name or vice-versa
    for emp_key, emp in emp_lookup['exact'].items():
        if key in emp_key or emp_key in key:
            return emp, 'substring'

    # 3. Close fuzzy match (cutoff 0.70)
    candidates = list(emp_lookup['exact'].keys())
    matches = difflib.get_close_matches(key, candidates, n=1, cutoff=0.70)
    if matches:
        return emp_lookup['exact'][matches[0]], f'fuzzy≈"{matches[0]}"'

    return None, 'not_found'


class Command(BaseCommand):
    help = (
        'STEP 1 — Convert the SdS salary sheet of the school accounts Excel '
        'into a CSV ready for bulk-import at /employees/payroll/bulk-import/'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel', required=True,
            help='Full path to the .xlsm / .xlsx accounts workbook',
        )
        parser.add_argument(
            '--output', default='datamigration/convertedcsvs/payroll_import.csv',
            help='Path for the output CSV file (default: datamigration/convertedcsvs/payroll_import.csv)',
        )
        parser.add_argument(
            '--sheet', default='SdS',
            help='Sheet name to read (default: SdS)',
        )
        parser.add_argument(
            '--session', default='2024-2025',
            help='Fallback session string when the row has no session column '
                 '(default: 2024-2025)',
        )
        parser.add_argument(
            '--skip-zero', action='store_true',
            help='Skip months where the salary amount is exactly 0',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Print what would be written without creating the output file',
        )

    def handle(self, *args, **options):
        excel_path = options['excel']
        output_path = options['output']
        sheet_name = options['sheet']
        fallback_session = options['session']
        skip_zero = options['skip_zero']
        dry_run = options['dry_run']

        # ── Validate inputs ────────────────────────────────────────────────
        if not os.path.exists(excel_path):
            raise CommandError(f'File not found: {excel_path}')

        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is not installed. Run: pip install openpyxl')

        fallback_start, fallback_end = _parse_session_years(fallback_session)
        if not fallback_start:
            raise CommandError(f'Invalid --session value: "{fallback_session}". Use format YYYY-YYYY.')

        # ── Load Excel sheet ───────────────────────────────────────────────
        self.stdout.write(f'Reading: {excel_path}  (sheet: {sheet_name})')
        wb = openpyxl.load_workbook(excel_path, read_only=True, keep_vba=True)
        if sheet_name not in wb.sheetnames:
            raise CommandError(
                f'Sheet "{sheet_name}" not found.\nAvailable sheets: {wb.sheetnames}'
            )

        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(min_row=1, max_col=25, values_only=True))
        if not raw_rows:
            raise CommandError('Sheet is empty.')

        # ── Parse header ───────────────────────────────────────────────────
        header = raw_rows[0]
        col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

        required_cols = {'Name'}
        missing = required_cols - set(col.keys())
        if missing:
            raise CommandError(f'Required column(s) missing from header: {missing}')

        month_cols = {m: col[m] for m in MONTH_ORDER if m in col}
        if not month_cols:
            raise CommandError('No month columns (April…March) found in header row.')

        session_col  = col.get('Session')
        old_dues_col = col.get('Old Dues')

        self.stdout.write(
            f'Month columns detected: {list(month_cols.keys())}\n'
            f'Old Dues column: {"yes" if old_dues_col is not None else "not found"}'
        )

        # ── Build employee lookup ──────────────────────────────────────────
        all_employees = list(Employee.objects.all())
        exact_lookup = {}
        for emp in all_employees:
            exact_lookup[emp.name.lower().strip()] = emp
            if emp.display_name:
                exact_lookup[emp.display_name.lower().strip()] = emp
        emp_lookup = {'exact': exact_lookup}

        self.stdout.write(f'Employees in database: {len(all_employees)}\n')

        # ── Process rows ───────────────────────────────────────────────────
        output_rows = []
        unmatched = []
        skipped_formula = []

        for row_num, row in enumerate(raw_rows[1:], start=2):
            name = row[col['Name']] if col['Name'] < len(row) else None

            # Skip blank, formula or note rows
            if not name or not isinstance(name, str):
                continue
            name = name.strip()
            if not name or name.startswith('=') or name[0].isdigit():
                skipped_formula.append((row_num, name))
                continue
            # Skip totals / annotation rows
            if any(kw in name.lower() for kw in ('total', 'sum', 'added for', '---')):
                skipped_formula.append((row_num, name))
                continue

            # Determine session years
            session_val = row[session_col] if (session_col is not None and session_col < len(row)) else None
            start_year, end_year = _parse_session_years(session_val)
            if not start_year:
                start_year, end_year = fallback_start, fallback_end

            # Old dues for this employee (carry-over from previous sessions)
            raw_old_dues = row[old_dues_col] if (old_dues_col is not None and old_dues_col < len(row)) else None
            try:
                emp_old_dues = float(raw_old_dues) if raw_old_dues is not None else 0.0
            except (ValueError, TypeError):
                emp_old_dues = 0.0
            # Only import positive old dues (negative means overpaid in prev session)
            emp_old_dues = max(emp_old_dues, 0.0)

            # Match to DB employee
            emp, match_type = _best_name_match(name, emp_lookup)
            if emp is None:
                unmatched.append((row_num, name))
                continue

            if match_type != 'exact':
                self.stdout.write(
                    self.style.WARNING(
                        f'  Row {row_num}: "{name}" matched to "{emp.name}" ({match_type})'
                    )
                )

            # Build one CSV row per month that has an entry
            old_dues_assigned = False
            for month_name in MONTH_ORDER:
                if month_name not in month_cols:
                    continue
                col_idx = month_cols[month_name]
                amount = row[col_idx] if col_idx < len(row) else None

                if amount is None:
                    continue            # No entry for this month — skip entirely
                try:
                    amount = float(amount)
                except (ValueError, TypeError):
                    continue

                if skip_zero and amount == 0.0:
                    continue

                month_str = _month_str(month_name, start_year, end_year)

                # Attach old dues to the first valid month for this employee
                entry_old_dues = 0.0
                if not old_dues_assigned and emp_old_dues > 0:
                    entry_old_dues = emp_old_dues
                    old_dues_assigned = True

                output_rows.append({
                    'Emp_ID':            emp.emp_no,
                    'Month':             month_str,
                    'Payable_Salary':    round(amount, 2),
                    'Old_Dues':          round(entry_old_dues, 2),
                    'Other_Amount':      0,
                    'Note':              '',
                    'Manual_Work_Days':  '',
                    'Manual_Leave_Days': '',
                })

        # ── Report ─────────────────────────────────────────────────────────
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'Rows to be written: {len(output_rows)}'))

        if skipped_formula:
            self.stdout.write(f'Skipped formula/note rows: {len(skipped_formula)}')

        if unmatched:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR(
                f'{len(unmatched)} employee name(s) could NOT be matched to a DB record:'
            ))
            self.stdout.write(self.style.ERROR(
                '  → Add these employees first via /employees/ or fix the name mismatch.'
            ))
            for row_num, name in unmatched:
                # Suggest closest DB name
                candidates = list(emp_lookup['exact'].keys())
                suggestions = difflib.get_close_matches(name.lower(), candidates, n=2, cutoff=0.50)
                hint = f'  (closest: {suggestions})' if suggestions else ''
                self.stdout.write(self.style.WARNING(f'  Row {row_num}: "{name}"{hint}'))

        if dry_run:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('DRY RUN — no file written. Remove --dry-run to save.'))
            self._print_preview(output_rows[:10])
            return

        # ── Write CSV ──────────────────────────────────────────────────────
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
            writer.writeheader()
            writer.writerows(output_rows)

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'✓ CSV written to: {os.path.abspath(output_path)}'))
        self.stdout.write(
            '\nSTEP 2 — Upload this CSV at:\n'
            '  http://127.0.0.1:8000/employees/payroll/bulk-import/\n'
            '  Use "Skip duplicates" on first run, "Update duplicates" to overwrite.'
        )

    def _print_preview(self, rows):
        if not rows:
            return
        self.stdout.write('\nPreview (first 10 rows):')
        header = '  '.join(f'{k:<20}' for k in OUTPUT_FIELDS)
        self.stdout.write(header)
        self.stdout.write('-' * len(header))
        for r in rows:
            line = '  '.join(f'{str(r[k]):<20}' for k in OUTPUT_FIELDS)
            self.stdout.write(line)
