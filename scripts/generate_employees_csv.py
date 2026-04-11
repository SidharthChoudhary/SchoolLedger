"""
Generate a unified employee CSV from both 2023-24 and 2024-25 Excel files.

Usage:
    python scripts/generate_employees_csv.py

Output:
    datamigration/convertedcsvs/employees_unified.csv

Rules:
- 2024-25 SdS sheet provides canonical names for staff active in 2024-25.
- 2023-24 StaticDataStaff provides additional staff who had already left by 2024-25.
- Junk rows (numeric Post, no session, note rows) are excluded.
- Near-duplicate names between years (e.g. "Munshi Driving" vs "Munsi Driver")
  are detected via fuzzy matching and merged, keeping the 2024-25 canonical name.
"""

import csv
import difflib
import os
import sys
import warnings

warnings.filterwarnings('ignore')

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl is required: pip install openpyxl')

XLSX_2324 = 'datamigration/accountsExcelFiles/DPS 2023-24_Accounts _updated Feb25.xlsm'
XLSX_2425 = 'datamigration/accountsExcelFiles/DPS 2024-25_Accounts_updated_22Jan_26_RSK.xlsm'
OUT = 'datamigration/convertedcsvs/employees_unified.csv'

FIELDS = [
    'Emp_No', 'Name', 'DOB', 'Contact_Number', 'Gender', 'Qualification',
    'Address', 'Experience_Years', 'Previous_Institute', 'Post', 'Role',
    'Role_Detail', 'Joining_Date', 'Base_Salary_Per_Month', 'Status', 'Leaves_Entitled',
]

START_EMP_NO = 1001


def _fmt_date(v):
    if v and hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return ''


def _is_valid_name(name):
    """Return True if name looks like a real employee name (not a formula/note row)."""
    if not name or not isinstance(name, str):
        return False
    name = name.strip()
    if not name or name[0].isdigit():
        return False
    n = name.lower()
    # Filter totals and annotation rows
    if n in ('total', 'sum', 'subtotal'):
        return False
    if n.startswith('total ') or 'added for' in n:
        return False
    # Filter all-lowercase names that look like notes (e.g. "rsk old dues")
    if name == name.lower() and ' ' in name:
        return False
    return True


def _load_sds_2425(path):
    """Return {name: {Post, Joining_Date, Base_Salary_Per_Month, Status}} from SdS sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True)
    ws = wb['SdS']
    rows = list(ws.iter_rows(min_row=2, max_col=7, values_only=True))
    result = {}
    for r in rows:
        if not _is_valid_name(r[0]):
            continue
        # Row: Name(0), Post(1), Joining Date(2), Base Salary(3), Session(4), Status(5), Old Dues(6)
        # Skip rows where Post is numeric (junk rows mixed into the sheet)
        if r[1] is not None and not isinstance(r[1], str):
            continue
        name = r[0].strip()
        result[name] = {
            'Post': str(r[1]).strip() if r[1] else '',
            'Joining_Date': _fmt_date(r[2]),
            'Base_Salary_Per_Month': r[3] if r[3] is not None else '',
            'Status': str(r[5]).strip() if r[5] else 'Active',
        }
    return result


def _load_staticdata_2324(path):
    """Return {name: {...}} from StaticDataStaff sheet, excluding junk rows."""
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True)
    ws = wb['StaticDataStaff']
    rows = list(ws.iter_rows(min_row=2, max_col=7, values_only=True))
    result = {}
    for r in rows:
        if not _is_valid_name(r[0]):
            continue
        # Row: Name(0), Post(1), Joining Date(2), Base Salary(3), Session(4), Status(5), Old Dues(6)
        # Must have a valid session string — filters out stray calculation rows
        session = r[4]
        if session and not isinstance(session, str):
            continue  # numeric in session column → junk row
        post = r[1]
        if post is not None and not isinstance(post, str):
            continue  # numeric in Post column → junk row
        # Must have at least a session or a Post to be a real employee row
        if not session and not post and r[3] is None:
            continue
        name = r[0].strip()
        result[name] = {
            'Post': str(post).strip() if post else '',
            'Joining_Date': _fmt_date(r[2]),
            'Base_Salary_Per_Month': r[3] if r[3] is not None else '',
            'Status': str(r[5]).strip() if r[5] else 'Active',
        }
    return result


def _fuzzy_match(name, candidates, cutoff=0.82):
    """Return the closest match from candidates, or None if below cutoff."""
    key = name.lower().strip()
    cand_keys = {c.lower().strip(): c for c in candidates}
    matches = difflib.get_close_matches(key, list(cand_keys.keys()), n=1, cutoff=cutoff)
    if matches:
        return cand_keys[matches[0]]
    return None


def main():
    print(f'Loading 2024-25: {XLSX_2425}')
    emp_2425 = _load_sds_2425(XLSX_2425)
    print(f'  {len(emp_2425)} employees from SdS')

    print(f'Loading 2023-24: {XLSX_2324}')
    emp_2324 = _load_staticdata_2324(XLSX_2324)
    print(f'  {len(emp_2324)} employees from StaticDataStaff')

    # Start with 2024-25 as the canonical list
    merged = list(emp_2425.items())
    names_2425 = list(emp_2425.keys())

    print('\nChecking 2023-24 employees not in 2024-25:')
    added = []
    skipped_fuzzy = []
    for name, data in emp_2324.items():
        # Exact match → already covered by 2024-25
        if name in emp_2425:
            continue
        # Fuzzy match → same person, different name spelling between years
        close = _fuzzy_match(name, names_2425)
        if close:
            skipped_fuzzy.append((name, close))
            print(f'  MERGED  "{name}" → "{close}" (same person, different spelling)')
            continue
        # Genuinely new employee from 2023-24
        merged.append((name, data))
        added.append(name)
        print(f'  ADDED   "{name}" [{data["Status"]}]')

    if skipped_fuzzy:
        print(f'\nNote: {len(skipped_fuzzy)} 2023-24 name(s) merged into 2024-25 canonical names.')
        print('The payroll import command will match them via fuzzy matching.')

    print(f'\nTotal employees: {len(merged)}')

    # Write CSV
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(FIELDS)
        for emp_no, (name, data) in enumerate(merged, start=START_EMP_NO):
            w.writerow([
                emp_no, name, '', '', '', '', '', '', '',
                data['Post'], '', '',
                data['Joining_Date'], data['Base_Salary_Per_Month'],
                data['Status'], 30,
            ])

    print(f'\nWritten to: {os.path.abspath(OUT)}')
    print('\nNEXT STEP:')
    print('  1. Review the CSV, adjust any names as needed.')
    print('  2. Import at: http://127.0.0.1:8000/employees/bulk-import/')
    print('  3. Then run the payroll CSV command:')
    print('     2023-24: python manage.py excel_to_payroll_csv \\')
    print('       --excel "datamigration/accountsExcelFiles/DPS 2023-24_Accounts _updated Feb25.xlsm" \\')
    print('       --sheet StaffStmtAll --header-row 4 --data-row 6 --session 2023-2024 \\')
    print('       --output datamigration/convertedcsvs/payroll_2023_24.csv --skip-zero')
    print('     2024-25: python manage.py excel_to_payroll_csv \\')
    print('       --excel "datamigration/accountsExcelFiles/DPS 2024-25_Accounts_updated_22Jan_26_RSK.xlsm" \\')
    print('       --output datamigration/convertedcsvs/payroll_2024_25.csv --skip-zero')


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    main()
